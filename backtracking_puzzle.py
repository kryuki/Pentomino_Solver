import sys
import time
sys.setrecursionlimit(10**9)
pieces = [None] * 13
pieces[0] = [None] * 1
pieces[0][0] = [[0, 0], [0, 2], [1, 0], [1, 1], [1, 2], (1, 6, 0, 2)]
pieces[1] = [None] * 4
pieces[1][0] = [[0, 0], [0, 1], [1, -1], [1, 0], [2, -1], (0, 5, 1, 6)]
pieces[1][1] = [[0, 0], [1, 0], [-1, -1], [0, -1], [-1, -2], (1, 6, 2, 7)]
pieces[1][2] = [[0, 0], [0, -1], [-1, 1], [-1, 0], [-2, 1], (2, 7, 1, 6)]
pieces[1][3] = [[0, 0], [-1, 0], [1, 1], [0, 1], [1, 2], (1, 6, 0, 5)]
pieces[2] = [None] * 2
pieces[2][0] = [[0, 0], [1, 0], [2, 0], [3, 0], [4, 0], (0, 3, 0, 7)]
pieces[2][1] = [[0, 0], [0, -1], [0, -2], [0, -3], [0, -4], (4, 7, 0, 7)]
pieces[3] = [None] * 1
pieces[3][0] = [[0, 0], [0, 1], [1, 0], [1, 1], (0, 6, 0, 6)]
pieces[4] = [None] * 8
pieces[4][0] = [[0, 0], [1, -1], [1, 0], [2, 0], [3, 0], (0, 4, 1, 7)]
pieces[4][1] = [[0, 0], [-1, -1], [0, -1], [0, -2], [0, -3], (1, 7, 3, 7)]
pieces[4][2] = [[0, 0], [-1, 1], [-1, 0], [-2, 0], [-3, 0], (3, 7, 0, 6)]
pieces[4][3] = [[0, 0], [1, 1], [0, 1], [0, 2], [0, 3], (0, 6, 0, 4)]
pieces[4][4] = [[0, 0], [1, 0], [1, 1], [2, 0], [3, 0], (0, 4, 0, 6)]
pieces[4][5] = [[0, 0], [0, -1], [0, -2], [0, -3], [1, -1], (0, 6, 3, 7)]
pieces[4][6] = [[0, 0], [-1, 0], [-2, 0], [-3, 0], [-1, -1], (3, 7, 1, 7)]
pieces[4][7] = [[0, 0], [0, 1], [0, 2], [0, 3], [-1, 1], (1, 7, 0, 4)]
pieces[5] = [None] * 1
pieces[5][0] = [[0, 0], [1, -1], [1, 0], [1, 1], [2, 0], (0, 5, 1, 6)]
pieces[6] = [None] * 4
pieces[6][0] = [[0, 0], [0, 1], [0, 2], [1, 2], [2, 2], (0, 5, 0, 5)]
pieces[6][1] = [[0, 0], [1, 0], [2, 0], [2, -1], [2, -2], (0, 5, 2, 7)]
pieces[6][2] = [[0, 0], [0, -1], [0, -2], [-1, -2], [-2, -2], (2, 7, 2, 7)]
pieces[6][3] = [[0, 0], [-1, 0], [-2, 0], [-2, 1], [-2, 2], (2, 7, 0, 5)]
pieces[7] = [None] * 4
pieces[7][0] = [[0, 0], [1, 0], [2, -1], [2, 0], [2, 1], (0, 5, 1, 6)]
pieces[7][1] = [[0, 0], [0, -1], [-1, -2], [0, -2], [1, -2], (1, 6, 2, 7)]
pieces[7][2] = [[0, 0], [-1, 0], [-2, 1], [-2, 0], [-2, -1], (2, 7, 1, 6)]
pieces[7][3] = [[0, 0], [0, 1], [1, 2], [0, 2], [-1, 2], (1, 6, 0, 5)]
pieces[8] = [None] * 8
pieces[8][0] = [[0, 0], [1, 0], [1, 1], [1, 2], [1, 3], (0, 6, 0, 4)]
pieces[8][1] = [[0, 0], [0, -1], [1, -1], [2, -1], [3, -1], (0, 4, 1, 7)]
pieces[8][2] = [[0, 0], [-1, 0], [-1, -1], [-1, -2], [-1, -3], (1, 7, 3, 7)]
pieces[8][3] = [[0, 0], [0, 1], [-1, 1], [-2, 1], [-3, 1], (3, 7, 0, 6)]
pieces[8][4] = [[0, 0], [-1, 0], [-1, 1], [-1, 2], [-1, 3], (1, 7, 0, 4)]
pieces[8][5] = [[0, 0], [0, 1], [1, 1], [2, 1], [3, 1], (0, 4, 0, 6)]
pieces[8][6] = [[0, 0], [1, 0], [1, -1], [1, -2], [1, -3], (0, 6, 3, 7)]
pieces[8][7] = [[0, 0], [0, -1], [-1, -1], [-2, -1], [-3, -1], (3, 7, 1, 7)]
pieces[9] = [None] * 8
pieces[9][0] = [[0, 0], [1, 0], [2, 0], [2, 1], [3, 1], (0, 4, 0, 6)]
pieces[9][1] = [[0, 0], [0, -1], [0, -2], [1, -2], [1, -3], (1, 7, 3, 7)]
pieces[9][2] = [[0, 0], [-1, 0], [-2, 0], [-2, -1], [-3, -1], (3, 7, 0, 6)]
pieces[9][3] = [[0, 0], [0, 1], [0, 2], [-1, 2], [-1, 3], (0, 6, 0, 4)]
pieces[9][4] = [[0, 0], [1, 0], [2, 0], [2, -1], [3, -1], (0, 4, 1, 7)]
pieces[9][5] = [[0, 0], [0, -1], [0, -2], [-1, -2], [-1, -3], (1, 7, 3, 7)]
pieces[9][6] = [[0, 0], [-1, 0], [-2, 0], [-2, 1], [-3, 1], (3, 7, 0, 6)]
pieces[9][7] = [[0, 0], [0, 1], [0, 2], [1, 2], [1, 3], (0, 6, 0, 4)]
pieces[10] = [None] * 4
pieces[10][0] = [[0, 0], [1, 0], [1, 1], [1, 2], [2, 2], (0, 5, 0, 5)]
pieces[10][1] = [[0, 0], [0, -1], [1, -1], [2, -1], [2, -2], (0, 5, 2, 7)]
pieces[10][2] = [[0, 0], [1, 0], [1, -1], [1, -2], [2, -2], (0, 5, 2, 7)]
pieces[10][3] = [[0, 0], [0, -1], [-1, -1], [-2, -1], [-2, -2], (2, 7, 2, 7)]
pieces[11] = [None] * 8
pieces[11][0] = [[0, 0], [0, 1], [1, 0], [1, 1], [1, 2], (0, 6, 0, 5)]
pieces[11][1] = [[0, 0], [1, 0], [0, -1], [1, -1], [2, -1], (0, 5, 1, 7)]
pieces[11][2] = [[0, 0], [0, -1], [-1, 0], [-1, -1], [-1, -2], (1, 7, 2, 7)]
pieces[11][3] = [[0, 0], [-1, 0], [0, 1], [-1, 1], [-2, 1], (2, 7, 0, 6)]
pieces[11][4] = [[0, 0], [0, 1], [0, 2], [1, 0], [1, 1], (0, 6, 0, 5)]
pieces[11][5] = [[0, 0], [1, 0], [2, 0], [0, -1], [1, -1], (0, 5, 1, 7)]
pieces[11][6] = [[0, 0], [0, -1], [0, -2], [-1, 0], [-1, -1], (1, 7, 2, 7)]
pieces[11][7] = [[0, 0], [-1, 0], [-2, 0], [0, 1], [-1, 1], (2, 7, 0, 6)]
pieces[12] = [None] * 8
pieces[12][0] = [[0, 0], [1, -1], [1, 0], [2, 0], [2, 1], (0, 5, 1, 6)]
pieces[12][1] = [[0, 0], [-1, -1], [0, -1], [0, -2], [1, -2], (1, 6, 2, 7)]
pieces[12][2] = [[0, 0], [-1, 1], [-1, 0], [-2, 0], [-2, -1], (2, 7, 1, 6)]
pieces[12][3] = [[0, 0], [1, 1], [0, 1], [0, 2], [-1, 2], (1, 6, 0, 5)]
pieces[12][4] = [[0, 0], [1, 0], [1, 1], [1, 2], [2, 1], (0, 5, 0, 5)]
pieces[12][5] = [[0, 0], [0, -1], [1, -1], [2, -1], [1, -2], (0, 5, 2, 7)]
pieces[12][6] = [[0, 0], [-1, 0], [-1, -1], [-1, -2], [-2, -1], (2, 7, 2, 7)]
pieces[12][7] = [[0, 0], [0, 1], [-1, 1], [-2, 1], [-1, 2], (2, 7, 0, 5)]

class UnionFind:
    def __init__(self, num):
        self.V = num
        self.parent = list(range(num))
        self.rank = [0 for _ in range(num)]
    def find_set(self, u):
        if u != self.parent[u]:
            self.parent[u] = self.find_set(self.parent[u])
        return self.parent[u]
    def union(self, set1, set2):
        if set1 == set2:
            return
        rank1, rank2 = self.rank[set1], self.rank[set2]
        if rank1 > rank2:
            self.parent[set2] = set1
        else:
            self.parent[set1] = set2
            if rank1 == rank2:
                self.rank[set2] += 1

# # for creating pieces
# def rotate_piece(piece):
#     #set the 0th pos in the piece at the origin
#     diff_i, diff_j = piece[0]
#     piece = [[i - diff_i, j - diff_j] for i, j in piece]

#     new_piece = []
#     for dir_i, dir_j in piece:
#         new_i, new_j = dir_j, -dir_i
#         new_piece.append([new_i, new_j])
    
#     new_diff_i, new_diff_j = new_piece[0]

#     return [[i - new_diff_i, j - new_diff_j] for i, j in new_piece]


# for i in range(12, 13):
#     print('piece index:', (i))
#     piece = pieces[i][4][:-1]
#     for _ in range(4):
#         print(piece)
#         piece = rotate_piece(piece)

grid = [['0' for _ in range(8)] for _ in range(8)]
# piece_0 = [[0, 0], [1, 0], [1, 1]]
# piece_1 = [[0, 0], [1, 0], [1, 1], [1, 2]]
# piece_2 = [[2, 0], [2, 1]]
# pieces = [piece_2, piece_1, piece_0]
# grid = [[0 for _ in range(3)] for _ in range(3)]

def in_range(i, j):
    return 0 <= i < 8 and 0 <= j < 8

def can_fit(grid, i, j, piece):
    for piece_pos_i, piece_pos_j in piece:
        grid_pos_i = piece_pos_i + i
        grid_pos_j = piece_pos_j + j
        if not in_range(grid_pos_i, grid_pos_j) or grid[grid_pos_i][grid_pos_j] != '0':
            return False
    return True

def fit_piece(grid, i, j, piece, idx):
    for piece_pos_i, piece_pos_j in piece:
        grid_pos_i = piece_pos_i + i
        grid_pos_j = piece_pos_j + j
        grid[grid_pos_i][grid_pos_j] = chr(ord('a') + idx)
    return grid

def remove_piece(grid, i, j, piece):
    for piece_pos_i, piece_pos_j in piece:
        grid_pos_i = piece_pos_i + i
        grid_pos_j = piece_pos_j + j
        grid[grid_pos_i][grid_pos_j] = '0'
    return grid

from collections import defaultdict
def sanity_check(grid):
    def get_idx(i, j):
        return i * 8 + j
    #union '0' s
    uf = UnionFind(64)
    moves = ((0, -1), (-1, 0))
    for i in range(8):
        for j in range(8):
            if grid[i][j] != '0':
                continue
            for move_i, move_j in moves:
                new_i, new_j = i + move_i, j + move_j
                if in_range(new_i, new_j) and grid[new_i][new_j] == '0':
                    uf.union(uf.find_set(get_idx(i, j)), uf.find_set(get_idx(new_i, new_j)))
    #get '0' islands
    islands = defaultdict(int)
    for i in range(8):
        for j in range(8):
            if grid[i][j] != '0':
                continue
            id = uf.find_set(get_idx(i, j))
            islands[id] += 1
    
    cnt4 = 0
    for val in islands.values():
        if val % 5 == 4:
            cnt4 += 1
        if val % 5 != 0 and val % 5 != 4:
            return False
    return cnt4 <= 1

import copy
res = []
cnt = 0

def make_tuple(grid):
    return tuple([tuple(line) for line in grid])

def backtrack(grid, idx):
    global cnt
    
    if idx == 13:
        # print(grid)
        # if grid not in res:
        cnt += 1
        # if grid in res:
        #     print("aloha!")
        print(cnt)
        res.append(copy.deepcopy(grid))
        return
    if not sanity_check(grid):
        return
    # if len(res) >= 11:
    #     return
    piece_candidates = pieces[idx]
    for piece in piece_candidates: #piece = [[0, 0], [0, 1], [1, -1], [1, 0], [2, -1], ((0, 5), (1, 6))]
        piece, move_range = piece[:-1], piece[-1]
        row_from, row_to, col_from, col_to = move_range
        for i in range(row_from, row_to + 1):
            for j in range(col_from, col_to + 1):
                if grid[i][j] != '0':
                    continue
                if can_fit(grid, i, j, piece):
                    grid = fit_piece(grid, i, j, piece, idx)
                    backtrack(grid, idx + 1)
                    grid = remove_piece(grid, i, j, piece)
    return

start = time.time()
backtrack(grid, 0)
# print(res)
# print(len(res))
print(time.time() - start)

#paint excel
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles.fonts import Font
from openpyxl.styles import borders
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.fills import PatternFill

wb = load_workbook("C:/Users/Ryuki/Desktop/sample.xlsx")
ws = wb["Sheet1"]
colors = ['EC6110', 
          'FFF100',
          '07AF7B',
          '3169B3',
          'F5A101',
          '68C8F2',
          'EF908A',
          '8A3B2C',
          'A53D92',
          'FFFFFF',
          'DEE2E5',
          '7D818D',
          '0A0000']

for num, grid in enumerate(res):
    r, c = num // 10, num % 10
    start_r, start_c = 9 * r, 9 * c
    for i in range(8):
        for j in range(8):
            color = colors[ord(grid[i][j]) - 97]
            ws.cell(row=start_r + i + 1, column=start_c + j + 1).fill = PatternFill(patternType='solid', start_color=color, end_color=color)


wb.save("C:/Users/Ryuki/Desktop/sample_complete.xlsx")